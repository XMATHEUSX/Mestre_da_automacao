from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.common.exceptions import *
from selenium.webdriver.common.keys import Keys
import time
import random
import os
import openpyxl
from tqdm import tqdm


class CursoAutomacao:
    def __init__(self):
        chrome_options = Options()
        chrome_options.add_argument('--lang=pt-BR')
        chrome_options.add_argument('--disable-notifications')

        self.driver = webdriver.Chrome(
            executable_path=os.getcwd()+os.sep+'chromedriver.exe', options=chrome_options)
        self.wait = WebDriverWait(
            driver=self.driver,
            timeout=10,
            poll_frequency=1,
            ignored_exceptions=[
                NoSuchElementException,
                ElementNotVisibleException,
                ElementNotSelectableException]
        )
        self.email = 'matheusxavier2302@gmail.com'
        self.senha = 'Bacon@#02'
        self.links_anuncio = []
        self.Vendedor = []
        self.numero = []

    def Digitando_como_uma_pessoa(self, texto, elemento):
        for letter in texto:
            elemento.send_keys(letter)
            time.sleep(random.randint(1, 10)/30)

    def login(self):
        time.sleep(3)
        email_box = self.driver.find_element_by_xpath(
            '//input[@type="email"]')
        self.Digitando_como_uma_pessoa(self.email, email_box)
        time.sleep(3)
        senha = self.driver.find_element_by_xpath(
            '//input[@type="text"]')
        self.Digitando_como_uma_pessoa(self.senha, senha)
        btn_login = self.driver.find_element_by_class_name('sc-kGXeez.dXzFBw')
        btn_login.click()

    def Coleta_dados_página(self):

        self.titulos = self.driver.find_elements_by_xpath(
            (f'//h2[@class="fnmrjs-10 deEIZJ"]'))

        self.valores = self.driver.find_elements_by_xpath(
            (f'//p[@class="fnmrjs-16 jqSHIm"]'))

        self.local = self.driver.find_elements_by_xpath(
            (f'//p[@class="fnmrjs-13 hdwqVC"]'))

        self.links = self.driver.find_elements_by_xpath(
            (f'//a[@data-lurker-detail="list_id"]'))
        for link in self.links:
            self.links_anuncio.append(link.get_attribute("href"))

    def Criando_planilha(self):
        self.Planilhas = openpyxl.Workbook()
        self.Planilhas.create_sheet(f'Preços {self.pesquisa}')
        self.planilhas_precos = self.Planilhas[f'Preços {self.pesquisa}']
        self.planilhas_precos.cell(row=1, column=1, value='Título')
        self.planilhas_precos.cell(row=1, column=2, value='Preço')
        self.planilhas_precos.cell(row=1, column=3, value='Localização')
        self.planilhas_precos.cell(row=1, column=4, value='Link')
        self.planilhas_precos.cell(row=1, column=5, value='Vendendor')
        self.planilhas_precos.cell(row=1, column=6, value='Numero Vendendor')

    def Grava_dados_na_planilha(self):
        for indice in range(0, len(self.titulos)):
            nova_linha = [self.titulos[indice].text,
                          self.valores[indice].text, self.local[indice].text, self.links_anuncio[indice]]
            self.planilhas_precos.append(nova_linha)

    def proxima_pagina(self):
        try:
            self.wait.until(expected_conditions.element_to_be_clickable(
                (By.LINK_TEXT, "Próxima pagina"))).click()
            # self.driver.find_element_by_link_text("Próxima pagina").click()
            self.Pagina()
        except Exception:
            pass

    def Pagina(self):
        self.Coleta_dados_página()
        self.Grava_dados_na_planilha()
        self.proxima_pagina()

    def Definindo_a_pesquisa(self):
        self.pesquisa = input("O que você quer pesquisar ?")

    def progress_barill(self, done):
        print("\rProgress: [{0:50s}] {1:.1f}%".format(
            'I' * int(done*50), done*100), end='')

    def acessando_anuncio(self):
        done = 0
        for link in self.links_anuncio:
            self.driver.get(f'{link}')
            done += ((len(self.links_anuncio)) /
                     ((len(self.links_anuncio)*len(self.links_anuncio))/100))
            self.progress_barill(done/100)
            time.sleep(10)
            try:
                self.Vendedor.append(self.driver.find_element_by_xpath(
                    ('//*[starts-with(@class,"sc-ifAKCX sc-jhAzac")]')).text)
            except Exception:
                self.Vendedor.append("Não foi possivel encontrar o vendedor")
            try:
                self.numero_exite = self.driver.find_element_by_class_name(
                    'sc-jTzLTM.sc-bSbAYC.epjKbD')
                self.numero_exite.click()
                time.sleep(3)
            # colocar o wait
            # Sempre colocar . no lugar do espaço quando pesquisar o elemento com find element
                self.numero.append(self.driver.find_element_by_class_name(
                    'sc-jTzLTM.sc-jhaWeW.jSLnFS').text)
            except Exception:
                self.numero.append("Sem numero")

            time.sleep(2)

    def Grava_dados_na_planilha_anunciante(self):
        for num in range(0, len(self.Vendedor)):
            self.planilhas_precos.cell(row=(
                num+2), column=5, value=f'{self.Vendedor[num]}')
            self.planilhas_precos.cell(row=(
                num+2), column=6, value=f'{self.numero[num]}')
        self.Planilhas.save(
            f'Preços {self.pesquisa} Na Região De São Paulo.xlsx')

    def iniciar(self):
        # self.driver.get(
        #    f'https://sp.olx.com.br/sao-paulo-e-regiao/regiao-de-jundiai?q={self.pesquisa}')
        self.driver.get('https://conta.olx.com.br/acesso')
        self.login()
        time.sleep(15)
        self.Definindo_a_pesquisa()
        self.driver.get(
            f'https://sp.olx.com.br/sao-paulo-e-regiao/regiao-de-jundiai?q={self.pesquisa}')
        self.Criando_planilha()
        self.Pagina()
        self.acessando_anuncio()
        self.Grava_dados_na_planilha_anunciante()


curso = CursoAutomacao()
curso.iniciar()
