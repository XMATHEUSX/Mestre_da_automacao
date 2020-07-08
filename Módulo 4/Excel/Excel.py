import openpyxl
import os

Planilhas = openpyxl.load_workbook(
    'Módulo 4'+os.sep+'Excel'+os.sep + 'ExemploPlanilha.xlsx')


print(Planilhas.sheetnames)
Planilha_2015 = Planilhas['Registros 2015']
print(Planilha_2015['B11'].value)
Planilha_2015['B11'] = 'Falcon'
Planilhas.save('Módulo 4'+os.sep+'Excel'+os.sep+'ExemploPlanilhav2.xlsx')
for linha in Planilha_2015.iter_rows(min_row=2, max_row=10):
    for celula in linha:
        print(celula.value)

for coluna in Planilha_2015.iter_cols(min_col=2, max_col=2, min_row=2):
    for celula in coluna:
        print(celula.value)


# USANDO FORMULAS EXCEL
Planilhas_funcoes = openpyxl.Workbook()
Planilhas_funcoes.create_sheet("Preços de Canetas")
planilhas_canetas = Planilhas_funcoes['Preços de Canetas']
planilhas_canetas.cell(row=1, column=1, value='=SUM(1,1)')
planilhas_canetas.cell(row=2, column=1, value='=SUM(5,5)')
planilhas_canetas.cell(row=3, column=1, value='=SUM(10,10)')
planilhas_canetas.cell(row=4, column=1, value='=AVERAGE(A1,A3)')
planilhas_canetas.cell(row=5, column=1, value='=MIN(A1:A4)')
planilhas_canetas.cell(row=6, column=1, value='=MAX(A1,A3)')
Planilhas_funcoes.save('Módulo 4'+os.sep+'Excel'+os.sep+'funcoes.xlsx')
