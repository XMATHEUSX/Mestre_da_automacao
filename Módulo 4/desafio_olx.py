from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.common.exceptions import *
from selenium.webdriver.common.keys import Keys
import time
import random
import os
import openpyxl


class CursoAutomacao:
    def __init__(self):
        chrome_options = Options()
        chrome_options.add_argument('--lang=pt-BR')
        chrome_options.add_argument('--disable-notifications')

        self.driver = webdriver.Chrome(
            executable_path=os.getcwd()+os.sep+'chromedriver.exe', options=chrome_options)
        self.wait = WebDriverWait(
            driver=self.driver,
            timeout=10,
            poll_frequency=1,
            ignored_exceptions=[
                NoSuchElementException,
                ElementNotVisibleException,
                ElementNotSelectableException]
        )

    def Digitando_como_uma_pessoa(self, texto, elemento):
        for letter in texto:
            elemento.send_keys(letter)
            time.sleep(random.randint(1, 10)/30)
    

    def Definindo_a_pesquisa(self):
        self.pesquisa = input("O que você quer pesquisar ?")

    def Criando_planilha(self):
        self.Planilhas = openpyxl.Workbook()
        self.Planilhas.create_sheet(f'Preços {self.pesquisa}')
        self.planilhas_precos = self.Planilhas[f'Preços {self.pesquisa}']
        self.planilhas_precos.cell(row=1, column=1, value='Título')
        self.planilhas_precos.cell(row=1, column=2, value='Preço)')
        self.planilhas_precos.cell(row=1, column=3, value='Localização')

    def Coleta_dados_página(self):
        self.titulos = self.driver.find_elements_by_xpath(
            (f'//h2[@class="fnmrjs-10 deEIZJ"]'))

        self.valores = self.driver.find_elements_by_xpath(
            (f'//p[@class="fnmrjs-16 jqSHIm"]'))

        self.local = self.driver.find_elements_by_xpath(
            (f'//p[@class="fnmrjs-13 hdwqVC"]'))
        
        self.link = self.driver.current_url

    def Grava_dados_na_planilha(self):
        for indice in range(0, len(self.titulos)):
            nova_linha = [self.titulos[indice].text,
                          self.valores[indice].text, self.local[indice].text]
            self.planilhas_precos.append(nova_linha)
        self.Planilhas.save(
            'Módulo 4'+os.sep+f'Preços {self.pesquisa} Na Região De São Paulo.xlsx')

    def proxima_pagina(self):
        try:
            self.wait.until(expected_conditions.element_to_be_clickable(
                (By.LINK_TEXT, "Próxima pagina"))).click()
            # self.driver.find_element_by_link_text("Próxima pagina").click()
            self.Pagina()
        except Exception:
            pass

    def Pagina(self):
        self.Coleta_dados_página()
        self.Grava_dados_na_planilha()
        self.proxima_pagina()

    def iniciar(self):
        self.Definindo_a_pesquisa()
        self.driver.get(
            f'https://sp.olx.com.br/sao-paulo-e-regiao/regiao-de-jundiai?q={self.pesquisa}')
        self.Criando_planilha()
        self.Pagina()


curso = CursoAutomacao()
curso.iniciar()
